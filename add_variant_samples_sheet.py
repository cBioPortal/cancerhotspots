#!/usr/bin/env python3
"""Add per-variant sheets (SNV, INDEL) to hotspots_v3.xlsx.

The existing Sheet/Old tabs only cover the 164 'new' SNV hotspots from
Bandlamudi et al. 2026 and aggregate all variants at a residue into a single
row. The v2 XLS preserved per-variant detail (one row per ref_AA x variant_AA)
on separate SNV and INDEL sheets. This script restores that structure for
the v3 XLSX across the full hotspot list (v2 + v3).

Two sheets are written:
  - SNV_Variants   : all 1274 SNV residues (1110 v2 + 164 v3)
  - INDEL_Variants : all 55 INDEL residues (all v2-era; v3 added no new indels)

Each row has a Hotspot_Version column ('v2' or 'v3') so readers can tell
whose analysis the per-tumor-type counts came from: v2 rows reflect the
2017 dataset, v3 rows reflect the 2026 Bandlamudi analysis.

Source files:
  webapp/src/main/resources/data/v3_multi_type_residue.txt
  webapp/src/main/resources/data/v3_multi_type_variant_file.txt
    (variant file contains each row 3x; we dedupe on read)

v3-ness heuristic: the v3 residue file is built as [v2 rows (1165)] + [v3-only
rows (164)], in order. Rows after line 1165 are the new-in-v3 hotspots. This
matches the 164 (Hugo, Residue) pairs populated on the 'Hotspot_Residues' tab.
"""
import os
from collections import OrderedDict

import openpyxl
from openpyxl.styles import Font

RESIDUE_FILE = "webapp/src/main/resources/data/v3_multi_type_residue.txt"
VARIANT_FILE = "webapp/src/main/resources/data/v3_multi_type_variant_file.txt"
XLSX = "webapp/src/main/resources/static/files/hotspots_v3.xlsx"

# Rows 1..V2_ROW_COUNT are v2 hotspots; later rows are new in v3.
V2_ROW_COUNT = 1165

HEADERS = [
    "Hugo_Symbol",
    "Codon",
    "Codon_Position",
    "Reference_Amino_Acid",
    "Variant_Amino_Acid",
    "Mutation_Count",
    "Samples",
    "Hotspot_Version",
]


def load_residues():
    """Return list of (hugo, residue, amino_acid_position, kind, version)."""
    out = []
    with open(RESIDUE_FILE) as f:
        hdr = f.readline().rstrip("\n").split("\t")
        ci = {h: i for i, h in enumerate(hdr)}
        for i, line in enumerate(f, start=1):
            fs = line.rstrip("\n").split("\t")
            kind = "INDEL" if fs[ci["Indel_Size"]].strip() else "SNV"
            version = "v2" if i <= V2_ROW_COUNT else "v3"
            out.append(
                (
                    fs[ci["Hugo_Symbol"]],
                    fs[ci["Residue"]],
                    fs[ci["Amino_Acid_Position"]],
                    kind,
                    version,
                )
            )
    return out


def load_variants():
    """(hugo, residue) -> list[(ref, alt, composition, count)], deduped, stable order."""
    by_residue = OrderedDict()
    seen = set()
    with open(VARIANT_FILE) as f:
        hdr = f.readline().rstrip("\n").split("\t")
        ci = {h: i for i, h in enumerate(hdr)}
        for line in f:
            if line in seen:
                continue
            seen.add(line)
            fs = line.rstrip("\n").split("\t")
            key = (fs[ci["Hugo_Symbol"]], fs[ci["Residue"]])
            ref = fs[ci["Reference_Amino_Acid"]]
            alt = fs[ci["Variant_Amino_Acid"]]
            comp = fs[ci["Tumor_Type_Composition"]]
            count = sum(int(p.rsplit(":", 1)[1]) for p in comp.split("|") if ":" in p)
            by_residue.setdefault(key, []).append((ref, alt, comp, count))
    return by_residue


def write_sheet(wb, name, residues, variants):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws.append(HEADERS)
    for c in ws[1]:
        c.font = Font(bold=True)

    rows_written = 0
    for hugo, residue, pos, _kind, version in residues:
        vlist = variants.get((hugo, residue))
        if not vlist:
            raise KeyError(f"No variant rows for {(hugo, residue)}")
        for ref, alt, comp, count in sorted(vlist, key=lambda v: (-v[3], v[1])):
            ws.append([hugo, residue, pos, ref, alt, count, comp, version])
            rows_written += 1
    return rows_written


def main():
    os.chdir(os.path.dirname(os.path.abspath(__file__)))

    residues = load_residues()
    variants = load_variants()

    snv = [r for r in residues if r[3] == "SNV"]
    ind = [r for r in residues if r[3] == "INDEL"]

    wb = openpyxl.load_workbook(XLSX)

    # Remove the previous single combined sheet if present (renamed below).
    if "Variant_Samples" in wb.sheetnames:
        del wb["Variant_Samples"]

    snv_rows = write_sheet(wb, "SNV_Variants", snv, variants)
    ind_rows = write_sheet(wb, "INDEL_Variants", ind, variants)

    wb.save(XLSX)
    v2_snv = sum(1 for r in snv if r[4] == "v2")
    v3_snv = sum(1 for r in snv if r[4] == "v3")
    print(
        f"SNV_Variants   : {snv_rows} rows across {len(snv)} residues "
        f"(v2: {v2_snv}, v3: {v3_snv})"
    )
    print(
        f"INDEL_Variants : {ind_rows} rows across {len(ind)} residues "
        f"(all v2)"
    )


if __name__ == "__main__":
    main()
