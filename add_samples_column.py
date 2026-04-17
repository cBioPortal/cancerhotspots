#!/usr/bin/env python3
"""Add the 'Samples' column (tumor-type frequency breakdown) to hotspots_v3.xlsx.

Source data: webapp/src/main/resources/data/v3_multi_type_residue.txt
Target:      webapp/src/main/resources/static/files/hotspots_v3.xlsx

The v2 XLS download had a 'Samples' column formatted as
'type:count|type:count|...'. The v3 XLSX was missing it. This script
regenerates the v3 XLSX with that column appended to each sheet.
"""

import os

import openpyxl
from openpyxl.styles import Font

DATA_FILE = "webapp/src/main/resources/data/v3_multi_type_residue.txt"
XLSX = "webapp/src/main/resources/static/files/hotspots_v3.xlsx"


def build_lookup():
    """Return {(Hugo_Symbol, Residue): Tumor_Type_Composition}."""
    lookup = {}
    with open(DATA_FILE) as f:
        header = f.readline().rstrip("\n").split("\t")
        ci = {h: i for i, h in enumerate(header)}
        for line in f:
            fields = line.rstrip("\n").split("\t")
            key = (fields[ci["Hugo_Symbol"]], fields[ci["Residue"]])
            lookup[key] = fields[ci["Tumor_Type_Composition"]]
    return lookup


def main():
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    lookup = build_lookup()
    wb = openpyxl.load_workbook(XLSX)

    # Sheet: (Hugo_Symbol, Codon) are columns 1, 2; add 'Samples' after last real column
    sheet = wb["Sheet"]
    samples_col_sheet = 25  # after 'adj_pval_MSK_vs_TCGA' (col 24)
    sheet.cell(row=1, column=samples_col_sheet, value="Samples").font = Font(bold=True)

    # Map (Hugo, position) -> samples string, for Old-sheet lookup
    position_to_samples = {}

    for r in range(2, sheet.max_row + 1):
        hugo = sheet.cell(row=r, column=1).value
        codon = sheet.cell(row=r, column=2).value
        if hugo is None:
            continue
        samples = lookup.get((hugo, codon))
        if samples is None:
            raise KeyError(f"No composition for {(hugo, codon)}")
        sheet.cell(row=r, column=samples_col_sheet, value=samples)
        pos = sheet.cell(row=r, column=3).value
        position_to_samples[(hugo, int(pos))] = samples

    # Old: Codon_Position is col 2 (numeric). Append Samples matched by (Hugo, pos).
    old = wb["Old"]
    # Place right after last real column ('Is_hotspot_annotated_in_OncoKB (V3.14)?' at col 6).
    # Column 7 holds a stray truncated header 'Is_h' with no data; overwrite it.
    samples_col_old = 7
    old.cell(row=1, column=samples_col_old, value="Samples").font = Font(bold=True)

    for r in range(2, old.max_row + 1):
        hugo = old.cell(row=r, column=1).value
        pos = old.cell(row=r, column=2).value
        if hugo is None:
            # keep trailing empty row blank
            old.cell(row=r, column=samples_col_old, value=None)
            continue
        samples = position_to_samples.get((hugo, int(pos)))
        if samples is None:
            raise KeyError(f"No composition for {(hugo, pos)}")
        old.cell(row=r, column=samples_col_old, value=samples)

    wb.save(XLSX)
    print(f"Wrote {XLSX} with 'Samples' column on both sheets.")


if __name__ == "__main__":
    main()
